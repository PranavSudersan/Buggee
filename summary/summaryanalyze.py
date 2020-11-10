import matplotlib
from matplotlib import rc
import matplotlib.pyplot as plt
import time
from datetime import datetime
import os
import os.path
from tkinter import filedialog
import tkinter as tk
import ast
import openpyxl
import pandas as pd
from pandas.io.json import json_normalize
import numpy as np
from PyQt5.QtWidgets import QFileDialog
import seaborn as sns
import cv2
import pingouin as pg
import itertools

matplotlib.use('Qt5Agg')
# import random

class SummaryAnal:

    def __init__(self): #initialize
        self.statDf = {}
        # self.df_forcedata = None
        # self.figdict = None
##        self.eq_count = [1,1,1,1] #fitting equation counter for each subplot
        # self.eq_count = {}
        # self.eq_count["All"] = [1,1,1,1]
        # self.summary_filepath = ""

    def importSummary(self, summary_filepath, delim='\t'): #import summary data
        print("import")
        self.fig = None
##        self.eq_count = [1,1,1,1]
        # self.eq_count = {}
        # self.eq_count["All"] = [1,1,1,1]
        # root = tk.Tk()
        # root.withdraw()
        # if filepath == None:
        #     # self.summary_filepath =  filedialog.askopenfilename(
        #     #     title = "Select summary data file")
        #      self.summary_filepath, _ = QFileDialog.getOpenFileName(caption = 
        #                                                             "Select summary data file")
        # else:
        #     self.summary_filepath = filepath
        
        # if self.summary_filepath != "":
        summarydf = pd.read_table(summary_filepath,delimiter=delim)
        
        #include paths
        summarydf['File location'] = summary_filepath #include filepath in data
        summarydf['File name'] = os.path.splitext(os.path.basename(summary_filepath))[0] #include filename in data
        
        col_list = summarydf.columns
        self.unitDict = {}
        for col_name in col_list:
            unit = col_name.split('[')[-1].split(']')[0]
            if unit == col_name:
                self.unitDict[col_name] = ''
            else:
                col_clean = col_name.split('[')[0].strip()
                summarydf.rename(columns = {col_name : col_clean}, 
                                 inplace=True)
                self.unitDict[col_clean] = ' [$' + unit + '$]'
                
        summarydf.reset_index(inplace = True) # include index as a column
                
            # with open(self.summary_filepath, 'r', encoding = "utf_8") as f: #open summary data file
            #     x = f.read().splitlines()
            # area_max = [[float(i) for i in ast.literal_eval(y.split('\t')[0])] for y in x[1:]]
            # area_pulloff = [[float(i) for i in ast.literal_eval(y.split('\t')[1])] for y in x[1:]]
            # force_adhesion = [[float(i) for i in ast.literal_eval(y.split('\t')[2])] for y in x[1:]]
            # adh_preload = [[float(i) for i in ast.literal_eval(y.split('\t')[3])] for y in x[1:]]
            # contact_time = [float(y.split('\t')[4]) for y in x[1:]]
            # speed = [[float(i) for i in ast.literal_eval(y.split('\t')[5])] for y in x[1:]]
            # steps = [ast.literal_eval(y.split('\t')[6]) for y in x[1:]]
            # force_friction = [[float(i) for i in ast.literal_eval(y.split('\t')[7])] for y in x[1:]]
            # area_friction = [[float(i) for i in ast.literal_eval(y.split('\t')[8])] for y in x[1:]]
            # friction_preload = [[float(i) for i in ast.literal_eval(y.split('\t')[9])] for y in x[1:]]
            # msrmnt_num = [float(y.split('\t')[10]) for y in x[1:]]
            # msrmnt_ok = [y.split('\t')[11] for y in x[1:]]
            # roi_label = [ast.literal_eval(y.split('\t')[12]) for y in x[1:]]
            # speed_def = [(ast.literal_eval(y.split('\t')[13])) for y in x[1:]] #speed definitions
            # error_vert = [float(y.split('\t')[14]) for y in x[1:]]
            # error_lat = [float(y.split('\t')[15]) for y in x[1:]]
            # slideStep = [y.split('\t')[16] for y in x[1:]]
            # roiarea_max = [[float(i) for i in ast.literal_eval(y.split('\t')[17])] for y in x[1:]]
            # roiarea_pulloff = [[float(i) for i in ast.literal_eval(y.split('\t')[18])] for y in x[1:]]            
            # length_max = [[float(i) for i in ast.literal_eval(y.split('\t')[19])] for y in x[1:]]
            # length_pulloff = [[float(i) for i in ast.literal_eval(y.split('\t')[20])] for y in x[1:]] 
            # roilength_max = [[float(i) for i in ast.literal_eval(y.split('\t')[21])] for y in x[1:]]
            # roilength_pulloff = [[float(i) for i in ast.literal_eval(y.split('\t')[22])] for y in x[1:]]
            # ecc_pulloff = [[float(i) for i in ast.literal_eval(y.split('\t')[23])] for y in x[1:]]
            # contnum_pulloff = [[float(i) for i in ast.literal_eval(y.split('\t')[24])] for y in x[1:]]
            # area_residue = [[float(i) for i in ast.literal_eval(y.split('\t')[25])] for y in x[1:]]
            # slope_header = x[0].split('\t')[26] #check if data exists
            # slope = [float(y.split('\t')[26]) if slope_header[:5] == 'Slope' and y.split('\t')[26] != '' else None for y in x[1:]]
            # self.slope_unit = slope_header.split('[')[1].split(']')[0] if slope_header[:5] == 'Slope' else None
            # k_beam_header = x[0].split('\t')[27] #check if data exists
            # k_beam = [float(y.split('\t')[27]) if k_beam_header[:4] == 'Beam' else None for y in x[1:]]
            # error_k_beam = [float(y.split('\t')[28]) if k_beam_header[:4] == 'Beam' else None for y in x[1:]]
            # deform_init = [float(y.split('\t')[29]) if k_beam_header[:4] == 'Beam' else None for y in x[1:]]
            # deform_pulloff = [float(y.split('\t')[30]) if k_beam_header[:4] == 'Beam' else None for y in x[1:]]
            # energy_adh = [float(y.split('\t')[31]) if k_beam_header[:4] == 'Beam' else None for y in x[1:]]
            # bound_area = [[float(i) for i in ast.literal_eval(y.split('\t')[32])] \
            #               if k_beam_header[:4] == 'Beam' else [None]*len(ast.literal_eval(y.split('\t')[12])) for y in x[1:]]
            # bound_peri = [[float(i) for i in ast.literal_eval(y.split('\t')[33])] \
            #               if k_beam_header[:4] == 'Beam' else [None]*len(ast.literal_eval(y.split('\t')[12])) for y in x[1:]]
            # bound_len = [[float(i) for i in ast.literal_eval(y.split('\t')[34])] \
            #              if k_beam_header[:4] == 'Beam' else [None]*len(ast.literal_eval(y.split('\t')[12])) for y in x[1:]]
            # bound_wid = [[float(i) for i in ast.literal_eval(y.split('\t')[35])] \
            #              if k_beam_header[:4] == 'Beam' else [None]*len(ast.literal_eval(y.split('\t')[12])) for y in x[1:]]
            # area_unit = x[0].split('\t')[0][-5:-1]
            # rownum = len(area_max)
            
            
            # data_folderpath = os.path.dirname(
            #                     os.path.dirname(
            #                         os.path.dirname(
            #                             self.summary_filepath)))
            
            
            # summarydf['Data Folder'] = data_folderpath
            # self.unitDict['Data Folder'] = ''
#             #data to be split according to roi label
#             header_split_max = ["Adhesion_Force", "Adhesion_Preload",
#                                 "Friction_Force","Friction_Preload"] #take max in "All" (legacy)
#             header_split_avg = ["Pulloff_Median_Eccentricity"] #take mean in "All" (legacy)
#             header_split_add = ["Max_Area", "Pulloff_Area", "Friction_Area",
#                                 "ROI_Max_Area", "ROI_Pulloff_Area",
#                                 "Max_Length", "Pulloff_Length", "ROI_Max_Length",
#                                 "ROI_Pulloff_Length", "Pulloff_Contact_Number",
#                                 "Residue_Area", "Max_Bounding_Area", 
#                                 "Max_Bounding_Perimeter", "Max_Bounding_Length", 
#                                 "Max_Bounding_Width"] #take sum in "All" (legacy)
#             header_split = header_split_max + header_split_avg + header_split_add
#             data_split = [force_adhesion, adh_preload,
#                           force_friction, friction_preload,
#                           ecc_pulloff,
#                           area_max, area_pulloff, area_friction, 
#                           roiarea_max, roiarea_pulloff,
#                           length_max, length_pulloff, roilength_max,
#                           roilength_pulloff, contnum_pulloff,
#                           area_residue, bound_area, bound_peri,
#                           bound_len, bound_wid]

#             #data not to be split according to roi label            
#             header_nosplit = ["Measurement_Number", "Measurement_OK", "Contact_Time",
#                               "Steps","ROI_Labels", "Speed", "Speed_Definition",
#                               "Error_Vertical", "Error_Lateral",
#                               "Sliding_Step", "Area_Units", "Data_Folder", "Slope",
#                               "Beam_Spring_Constant","Error_Beam_Spring_Constant",
#                               "Initial_Deformation","Pulloff_Deformation","Adhesion_Energy"]
#             data_nosplit = [msrmnt_num, msrmnt_ok, contact_time,
#                             steps, roi_label, speed, speed_def,
#                             error_vert,error_lat,
#                             slideStep, [area_unit] * rownum,
#                             [data_folderpath] * rownum, slope,
#                             k_beam, error_k_beam, deform_init,
#                             deform_pulloff, energy_adh]
            
#             header_raw = header_nosplit + header_split 
#             data_raw = data_nosplit + data_split 
            
#             #define dictionaries for each step/roi and combine
#             header_dict = [a + "_Dict" for a in header_split]
#             data_dict = []
#             for j in range(len(header_split)):
#                 temp_dict = [dict(zip([header_split[j] + "_" + s for s in roi_label[i]],
#                                    data_split[j][i])) for i in range(len(roi_label))]
#                 data_dict.append(temp_dict)
            
#             header = header_raw + header_dict
#             datalist = data_raw + data_dict
            
#             datadict = dict(zip(header, datalist))
#             df_data = pd.DataFrame(datadict)

            
#             #split steps into columns and combine
#             df_speed_steps = json_normalize(df_data['Speed_Definition']) #split speed steps
#             df_all_data = [df_data, df_speed_steps]
#             for a in header_dict:
#                 df_temp = json_normalize(df_data[a])
#                 df_all_data.append(df_temp)

#             df_combined = pd.concat(df_all_data, join='outer', axis=1).fillna(np.nan)
#             df_combined.drop(header_dict, inplace=True, axis=1) #drop dictionary columns
#             df_good = df_combined[df_combined["Measurement_OK"] == "Y"]

# ##            self.steps_unique = set([a for b in steps_modif for a in b])
#             roi_label_unique = set([a for b in df_good["ROI_Labels"] for a in b])
# ##            self.speed_def_unique = set([a for b in df_good["Speed_Definition"] for a in b.keys()])

#             #reshape and combine roi data into new dataframe
#             header_nocomb = ["ROI Label", "Data_Folder",
#                              "Measurement_Number", "Measurement_OK",
#                              "Contact_Time", "Detachment Speed",
#                              "Attachment Speed", "Sliding Speed", "Sliding_Step",
#                              "Error_Vertical", "Error_Lateral", "Area_Units", "Slope",
#                              "Beam_Spring_Constant","Error_Beam_Spring_Constant",
#                              "Initial_Deformation","Pulloff_Deformation","Adhesion_Energy"]                      
#             header_comb = ["Adhesion_Force", "Adhesion_Preload",
#                           "Friction_Force","Friction_Preload",
#                           "Max_Area", "Pulloff_Area", "Friction_Area",
#                           "ROI_Max_Area", "ROI_Pulloff_Area",
#                           "Max_Length", "Pulloff_Length", "ROI_Max_Length",
#                           "ROI_Pulloff_Length", "Pulloff_Contact_Number",
#                           "Residue_Area", "Pulloff_Median_Eccentricity",
#                           "Max_Bounding_Area", "Max_Bounding_Perimeter", 
#                           "Max_Bounding_Length", "Max_Bounding_Width"]
#             header_all = header_nocomb + header_comb
#             self.df_forcedata = pd.DataFrame(columns = header_all)

            
#             for b in roi_label_unique:
#                 data_nocomb = [b] + [df_good[x] \
#                                      for x in header_nocomb \
#                                      if x not in ["ROI Label"]]
#                 data_comb = [df_good[x + "_" + b] for x in header_comb]

#                 df_nocomb = pd.DataFrame(dict(zip(header_nocomb, data_nocomb)))
#                 df_comb = pd.DataFrame(dict(zip(header_comb, data_comb)))
#                 df_joined = df_comb.join(df_nocomb)
#                 self.df_forcedata = self.df_forcedata.append(df_joined, ignore_index=True, sort=False)
#                 self.df_forcedata['Adhesion_Force'].replace('', np.nan, inplace=True)
#                 self.df_forcedata.dropna(subset=['Adhesion_Force'], inplace=True) #remove blanks

            #calculate additional data
            # self.df_forcedata['Adhesion_Stress'] = self.df_forcedata['Adhesion_Force']/self.df_forcedata['Pulloff_Area']
            # self.df_forcedata['Friction_Stress'] = self.df_forcedata['Friction_Force']/self.df_forcedata['Friction_Area']
            # self.df_forcedata['Normalized_Adhesion_Force'] = self.df_forcedata['Adhesion_Force']/self.df_forcedata['Max_Area']
            # self.df_forcedata['Normalized_Adhesion_Energy'] = self.df_forcedata['Adhesion_Energy']/self.df_forcedata['Max_Area']
            # self.df_forcedata['Date_of_Experiment'] =  self.df_forcedata['Data_Folder'].str.split(pat = "/").str[-1].str.slice(start=0, stop=9)
        #TODO: remove these, make variable definitions in dialog to calculate
        # summarydf['Adhesion Stress'] = summarydf['Pulloff Force']/summarydf['Pulloff Area']
        # self.unitDict['Adhesion Stress'] = ' [$' + self.extractUnit('Pulloff Force') + \
        #     '/' + self.extractUnit('Pulloff Area') + '$]'
        # summarydf['Friction Stress'] = summarydf['Friction Force']/summarydf['Friction Area']
        # self.unitDict['Friction Stress'] = ' [$' + self.extractUnit('Friction Force') + \
        #     '/' + self.extractUnit('Friction Area') + '$]'
        # summarydf['Normalized Adhesion Force'] = summarydf['Pulloff Force']/summarydf['Max Area']
        # self.unitDict['Normalized Adhesion Force'] = ' [$' + self.extractUnit('Pulloff Force') + \
        #     '/' + self.extractUnit('Max Area') + '$]'
        # summarydf['Normalized_Adhesion_Energy'] = summarydf['Adhesion Energy']/summarydf['Max Area']
        # self.unitDict['Normalized_Adhesion_Energy'] = ' [$' + self.extractUnit('Adhesion Energy') + \
        #     '/' + self.extractUnit('Max Area') + '$]'
        
        # summarydf['Date of Experiment'] =  summarydf['Data Folder'].str.split(pat = "/").str[-1].str.slice(start=0, stop=9)
        # self.unitDict['Date of Experiment'] = ''
        # self.summarydf = summarydf.copy()
        return summarydf
            # print(summarydf)
            # print(self.unitDict)
            
            # self.df_forcedata.reset_index(inplace = True, drop = True)
            # self.df_final = self.df_forcedata.copy()

##            #create combined "All" data by taking sum/mean/max among various roi
####            if len(self.roi_label_unique) > 1: #ROI names MUST NOT have "All" ot "Dict"!
##            columns_all = [a + "_All" for a in header_split]
##            self.df_forcedata = pd.concat([self.df_forcedata,
##                                           pd.DataFrame(columns=columns_all)], sort=False)
##            self.df_forcedata[columns_all] = self.df_forcedata[columns_all].fillna(0)
##
##            for a in self.roi_label_unique:
##                if a == 'All':
##                    print("Change ROI name 'All'")
##                    break
##                for i in range(len(columns_all)):
##                    if header_split[i] in header_split_add:
##                        self.df_forcedata[columns_all[i]] +=  self.df_forcedata[header_split[i] +
##                                                                                "_" + a].fillna(0)
##                    elif header_split[i] in header_split_max:
##                        clist1 = [header_split[i] + "_" + b for b in self.roi_label_unique]
##                        self.df_forcedata[columns_all[i]] = self.df_forcedata[clist1].max(axis=1)
##                    elif header_split[i] in header_split_avg:
##                        clist2 = [header_split[i] + "_" + b for b in self.roi_label_unique]
##                        self.df_forcedata[columns_all[i]] = self.df_forcedata[clist2].mean(axis=1)
##
##            self.roi_label_unique.update(["All"])
##            self.df_final.to_excel('E:/Work/Data/Summary/20200213/Sex/summary_temp_' +
##                                   str(random.randint(1, 90000)) + '.xlsx') #export as excel        

    def extractUnit(self, col):
        return self.unitDict[col].split('[')[-1].split(']')[0].replace('$', '')

    def filter_df(self, df, filter_dict): #filter df based on condition
        print(filter_dict)
        df_final = df.copy()
        for k in filter_dict.keys():
            col = filter_dict[k][0]
            cond = filter_dict[k][1]
            val = np.array([filter_dict[k][2]]).astype(df_final[col].dtype)[0]
            # if col in ["Weight","Temperature","Humidity","Contact_Angle-Water", 
            #            "Contact_Angle-Hexadecane","Measurement_Number","Contact_Time", 
            #            "Detachment Speed", "Attachment Speed", "Sliding Speed"]:
            #     val = float(filter_dict[k][2])
            # elif col in ["Folder_Name", "Species", "Sex", "Leg", "Pad","Medium",
            #                "Substrate","Label", "ROI Label","Sliding_Step"]:
            #     val = filter_dict[k][2]
            # elif col in ["Date"]:
            #     val = datetime.strptime(filter_dict[k][2], "%d/%m/%Y").date()
            if cond == 'equal to':
                print("equal condition")
                df_final = df_final[df_final[col] == val]
            elif cond == 'not equal to':
                df_final = df_final[df_final[col] != val]
            elif cond == 'greater than':
                df_final = df_final[df_final[col] > val]
                print(df_final[col].head())
                print("greater than", val)
            elif cond == 'less than':
                df_final = df_final[df_final[col] < val]
            elif cond == 'greater than or equal to':
                df_final = df_final[df_final[col] >= val]
            elif cond == 'less than or equal to':
                df_final = df_final[df_final[col] <= val]
        return df_final
            
    def get_units(self, var, df):
        if var in ["Adhesion_Force", "Adhesion_Preload",
                   "Friction_Force", "Friction_Preload"]: #force
            unit = ' $(μN)$'
        elif var in ["Max_Area", "Pulloff_Area",
                     "Friction_Area", "ROI_Max_Area",
                     "ROI_Pulloff_Area", "Max_Bounding_Area"]: #area
            unit = ' $(' + df["Area_Units"].iloc[0] + ')$'
        elif var in ["Max_Length", "Pulloff_Length",
                     "ROI_Max_Length", "ROI_Pulloff_Length",
                     "Max_Bounding_Perimeter", "Max_Bounding_Length", 
                     "Max_Bounding_Width"]: #length
            unit = ' $(' + df["Area_Units"].iloc[0][:-2] + ')$'
        elif var in ["Detachment Speed", "Attachment Speed",
                     "Sliding Speed"]: #speed
            unit = ' $(μm/s)$'
        elif var in ["Contact_Time"]: #time
            unit = ' $(s)$'
        elif var in ["Slope"]: #slope
            unit = self.slope_unit
        elif var in ["Adhesion_Stress", "Friction_Stress", "Normalized_Adhesion_Force"]:
            unit = ' $(μN' + '/' + df["Area_Units"].iloc[0] + ')$'
        elif var in ["Beam_Spring_Constant"]:
            unit = ' $(μN/μm)$'
        elif var in ["Initial_Deformation", "Pulloff_Deformation"]:
            unit = ' $(μm)$'
        elif var in ["Adhesion_Energy"]:
            unit = ' $(pJ)$'
        elif var in ["Normalized_Adhesion_Energy"]:
            unit = ' $(J/m^2)$'
        elif var in ["Contact_Angle-Water", "Contact_Angle-Hexadecane"]:
            unit = r' $(°)$'
        elif var in ["Temperature"]:
            unit = r' $(°C)$'
        elif var in ["Humidity"]:
            unit = ' $(%)$'
        elif var in ["Weight"]:
            unit = ' $(g)$'
        else:
            unit = ''
        return unit

    def get_errordata(self, var, df): #get errorbar data
        if var in ["Adhesion_Force", "Adhesion_Preload"]:
            error = df["Error_Vertical"]
        elif var in ["Friction_Force", "Friction_Preload"]:
            error = df["Error_Lateral"]
        elif var in ["Beam_Spring_Constant"]:
            error = df["Error_Beam_Spring_Constant"]
        else:
            error = None
        return error
    
    def plotSummary(self, df, paramDict):
        # if paramDict == None:
        #     paramDict = {'Plot type': 'scatter',
        #                  'X Variable': 'Measurement number',
        #                  'Y Variable': 'Adhesion Stress',
        #                  'Color Parameter': 'Detachment Speed',
        #                  'Row Parameter': None,
        #                  'Column Parameter': None,
        #                  'Style Parameter': 'ROI Label',
        #                  'Size Parameter': None,
        #                  'Show Markers': True,
        #                  'Opacity': 0.8,
        #                  'Title': 'Test PLot',
        #                  'Plot context': 'talk',
        #                  'Plot style': 'ticks',
        #                  'Color palette': None,
        #                  'Zero line': False,
        #                  'Legend outside': False,
        #                  'Legend location': 'center right',
        #                  'Legend columns': 1,
        #                  'X label rotate': 45}
        if self.fig != None:
            plt.close(self.fig.fig)
        
        x_var = paramDict['X Variable'].currentText() \
            if paramDict['X Variable'].currentText() != 'None' else None
        y_var = paramDict['Y Variable'].currentText() \
            if paramDict['Y Variable'].currentText() != 'None' else None
        hue_var = paramDict['Color Parameter'].currentText() \
            if paramDict['Color Parameter'].currentText() != 'None' else None
        row_var = paramDict['Row Parameter'].currentText() \
            if paramDict['Row Parameter'].currentText() != 'None' else None
        col_var = paramDict['Column Parameter'].currentText() \
            if paramDict['Column Parameter'].currentText() != 'None' else None
        style_var = paramDict['Style Parameter'].currentText() \
            if paramDict['Style Parameter'].currentText() != 'None' else None
        size_var = paramDict['Size Parameter'].currentText() \
            if paramDict['Size Parameter'].currentText() != 'None' else None
        color_pal = paramDict['Color palette'].currentText() \
            if paramDict['Color palette'].currentText() != 'None' else None

        
        #location dictionary
        locDict = {'best': 0,
                   'upper right': 1,
                   'upper left': 2,
                   'lower left': 3,
                   'lower right': 4,
                   'right': 5,
                   'center left': 6,
                   'center right': 7,
                   'lower center': 8,
                   'upper center': 9,
                   'center': 10}
            
        sns.set_theme(context = paramDict['Plot context'].currentText(),
                      style= paramDict['Plot style'].currentText(),
                      palette = color_pal,
                      font_scale = paramDict['Font scale'].value())
        
        # fig_size = (6,6) #figure size
        # x_param = "Distance"
        # x_label = "Distance, d/s"
        # y_param = "Force"
        # y_label = r"Force, $F/\gamma s$"
        # hue_param = "Model" #"Model"
        # hue_param_order = None #set to force_order to show multiple forces in color, else set to None for single force
        # style_param = 'Model'
        # style_param_order = None
        # col_param = "Contact Angle"
        # col_param_order = None
        #set to None if dont want to group and find Max
        # color_list = None#['g', 'b', 'r', 'c'] #corresponding colors
        fix_marker = 'o' if style_var == None else None #set to 'o' if style_param is None else set as None
        # line_styles = None#[(1, 0), (1, 1)]
        ##group_param = ['Model', 'Contact Angle', 'θ-fa', 'θ-fw',
        ##               'D_p/D_h', 'a_b', 'a_f', 'Fluid data file',
        ##               'Bubble data file'] 
        if paramDict['Plot type'].currentText() in ['line', 'scatter']:
            self.fig = sns.relplot(data= df,
                                   x = x_var,
                                   y = y_var,
                                   hue = hue_var,
                                   row = row_var,
                                   col = col_var,
                                   style = style_var,
                                   size = size_var,
                                   kind = paramDict['Plot type'].currentText(),
                                   markers = paramDict['Show Markers'].isChecked(),
                                   marker =  fix_marker)
        elif paramDict['Plot type'].currentText() in ["strip", "swarm", "box", 
                                                      "violin","boxen", "point", 
                                                      "bar", "count"]:
            self.fig = sns.catplot(data= df,
                                   x = x_var,
                                   y = y_var,
                                   hue = hue_var,
                                   row = row_var,
                                   col = col_var,
                                   kind = paramDict['Plot type'].currentText())
                                   #alpha = paramDict['Opacity'].value())            
        
        axes = self.fig.axes.flatten() #list of axes
        
        # ax.fig.set_size_inches(*fig_size)
        self.fig.set_axis_labels(x_var + self.unitDict[x_var],
                                 y_var + self.unitDict[y_var])
        
        if paramDict['X label rotate'].value() != 0:
            self.fig.set_xticklabels(rotation = paramDict['X label rotate'].value(),
                                     horizontalalignment = 'right')
        #rename subplot titles with unit
        for ax in axes:
            title = ax.get_title().split(' = ')
            if title[0] in self.unitDict.keys():
                ax.set_title(' = '.join(title) + 
                             self.unitDict[title[0]].replace('[', '').replace(']', ''))
        
        #show dashed zero line
        if paramDict['Zero line'].isChecked() == True:
            self.fig.map(plt.axhline, y=0, color=".7", dashes=(2, 1), zorder=0)

        self.fig.tight_layout(w_pad=0)
        
        leg = self.fig._legend
        if leg != None:
            handles, labels = axes[0].get_legend_handles_labels()
            leg_title = leg.get_title().get_text()
            
            #rename legend with units
            i= 0
            unit = ''
            for lab in labels:
                if leg_title == '':
                    if lab in self.unitDict.keys():
                        unit = self.unitDict[lab]
                        i += 1
                        continue
                else:
                    unit = self.unitDict[leg_title]
                unit = unit.replace('[', '').replace(']', '') if unit != '' else ''
                labels[i] = lab + unit
                i += 1
            
            leg.remove()        
            self.fig.fig.tight_layout()
            
            #redraw legend for better control
            leg = plt.legend(handles, labels, ncol = paramDict['Legend columns'].value(), 
                             framealpha = 0.8,
                             title = leg_title if leg_title != ''  else None)
            if paramDict['Legend outside'].isChecked() == True:
                leg_bbox = leg.get_tightbbox(self.fig.fig.canvas.get_renderer())
                x0, y0, w, h = leg_bbox.inverse_transformed(self.fig.fig.transFigure).bounds
                leg.set_bbox_to_anchor((-w, -h, 1 + 2 * w, 1 +  2 * h), 
                                       transform = self.fig.fig.transFigure)
            leg._loc = locDict[paramDict['Legend location'].currentText()]
        
        #despine axes
        if paramDict['Despine'].isChecked() == True:
            sns.despine(offset=10, trim=True)
        
        self.fig.fig.canvas.set_window_title(paramDict['Title'].text())
        self.fig.fig.canvas.mpl_connect("resize_event", self.previewPlot)
        
        # self.fig.fig.canvas.draw()
        
        # self.previewPlot()
        # self.fig.fig.show()
        # plt.show()

    #save plot to temporarory location and preview it. Necessary since
    #legends is outside view in plot show
    def previewPlot(self, event = None):        
        self.fig.fig.savefig("samples/_temp_plot.png", bbox_inches='tight', 
                             transparent = True, dpi = 50)
        temp_plot = cv2.imread("samples/_temp_plot.png")
        cv2.imshow("Plot Preview", temp_plot)
        
    #calculate statistical comparisons (ANOVA/t-test)
    def calculateStats(self, df, paramDict):
        ##reference: https://reneshbedre.github.io/blog/anova.html
        #IMPORTANT! requirements XlsxWriter==1.3.7, statsmodels==0.12.0
        
        # if get_stats == True:
            # from pingouin import anova, normality, homoscedasticity, pairwise_tukey
        ##    import statsmodels.api as sm
        ##    from statsmodels.formula.api import ols
        ##    # Ordinary Least Squares (OLS) model
        ##    # C(Label):C(Substrate) represent interaction term
        ##    model = ols('Adhesion_Force ~ C(Contact_type) + C(Substrate) + C(Contact_type):C(Substrate)',
        ##                data=exptData).fit()
        ##    anova_table = sm.stats.anova_lm(model, typ=2)
        ##    print('ANOVA:\n', anova_table)
        
        #remove spaces from column names
        df_clean = df.copy()
        col_list = df_clean.columns
        for col_name in col_list:
            col_clean = col_name.replace(' ', '_')
            df_clean.rename(columns = {col_name : col_clean},
                            inplace=True)
                
        x_var = paramDict['X Variable'].currentText().replace(' ', '_')
        y_var = paramDict['Y Variable'].currentText().replace(' ', '_')
        hue_var = paramDict['Color Parameter'].currentText().replace(' ', '_')
        row_var = paramDict['Row Parameter'].currentText().replace(' ', '_')
        col_var = paramDict['Column Parameter'].currentText().replace(' ', '_')
            
        group_vars = [a for a in [x_var, hue_var, row_var, col_var]  if a != 'None']
        
        #Two-way ANOVA test
        #TODO: also use welch_anova() for unequal variances
    ##    from pingouin import anova
        anovaDf = pg.anova(data = df_clean,
                           dv = y_var,
                           between = group_vars).round(3)
        print('ANOVA:\n', anovaDf)
    
        # Shapiro-Wilk test used to check the normal distribution of residuals
    
    ##    import scipy.stats as stats
    ##    w, pvalue = stats.shapiro(model.resid)
    ##    print('Shapiro-Wilk test:\n', w, pvalue)
    
    ##    from pingouin import normality
    #     norm_glass = pg.normality(exptData[exptData['Substrate'] == 'Glass'],
    #                            dv= y_var,
    #                            group='Contact_type',
    #                            method = 'shapiro').round(3)
    #     norm_glass['Substrate'] = 'Glass'
    #     print('Shapiro-Wilk test (glass):\n', norm_glass)
    #     norm_pfots = pg.normality(exptData[exptData['Substrate'] == 'PFOTS'],
    #                            dv='Adhesion_Force',
    #                            group='Contact_type',
    #                            method = 'shapiro').round(3)
    #     norm_pfots['Substrate'] = 'PFOTS'
    #     print('Shapiro-Wilk test (PFOTS):\n', norm_pfots)
    
    #     norm_test = norm_glass.append(norm_pfots)
        
    #     #Bartlett’s test to check the Homogeneity of variances
    #     ##import scipy.stats as stats
    #     ##w, pvalue = stats.bartlett(d['A'], d['B'], d['C'], d['D'])
    #     ##print(w, pvalue)
    
    #     #Levene's/Bartlett’s Test to check equality of variance (levene or bartlett)
    # ##    from pingouin import homoscedasticity
    #     lev_glass = pg.homoscedasticity(exptData[exptData['Substrate'] == 'Glass'],
    #                            dv=y_var,
    #                            group='Contact_type',
    #                            method = 'levene').round(3)
    #     lev_glass['Substrate'] = 'Glass'
    #     print('Levene Test (glass):\n', lev_glass)
    #     lev_pfots = pg.homoscedasticity(exptData[exptData['Substrate'] == 'PFOTS'],
    #                            dv='Adhesion_Force',
    #                            group='Contact_type',
    #                            method = 'levene').round(3)
    #     lev_pfots['Substrate'] = 'PFOTS'
    #     print('Levene Test (PFOTS):\n', lev_pfots)
    
    #     var_eq_test = lev_glass.append(lev_pfots)
    
        # perform multiple pairwise comparison (Tukey HSD)
        # for unbalanced (unequal sample size) data, pairwise_tukey uses Tukey-Kramer test
    ##    from pingouin import pairwise_tukey
        
        tukeyDf = pd.DataFrame()
        normDf = pd.DataFrame()
        vareqDf = pd.DataFrame()
        #TODO: ALSO USE pairwise_ameshowell() for unequal variances
        # fixed_params = ["Substrate", "Contact_type"]
        df_val_dict = {}
        for var in group_vars:
            df_val_dict[var] = df_clean[var].unique()
            
        for i in range(len(group_vars)):
            between_var = group_vars[i]
            var_fixed = [group_vars[a] for a in range(len(group_vars)) \
                         if a != i]
            if len(var_fixed) != 0:
                val_fixed = [df_val_dict[var] for var in var_fixed]
                #all possible combinations of fixed values
                fixed_comb = [a for a in itertools.product(*val_fixed)]                
                for comb in fixed_comb:
                    cond = [True]*df_clean.shape[0]
                    #df filter condition based on constant values of var_fixed combinations
                    for k in range(len(var_fixed)):
                        cond = (df_clean[var_fixed[k]] == comb[k]) & (cond)
                    #round numeric value    
                    comb = map(lambda a: round(a,3),comb) \
                        if comb[0].__class__.__name__ in ['float', 'float64'] \
                            else comb
                    # print([a for a in comb], var_fixed, 'comb')
                    # print(str(dict(zip(var_fixed, comb))))
                    fixed_param_string = str(dict(zip(var_fixed, comb))).\
                        replace('{','').replace('}','').replace("'",'')

                    # perform multiple pairwise comparison (Tukey HSD)
                    # for unbalanced (unequal sample size) data, pairwise_tukey uses Tukey-Kramer test                        
                    tukey_result = pg.pairwise_tukey(data=df_clean[cond],
                                                     dv= y_var,
                                                     between= between_var,
                                                     effsize = 'r')
                    tukey_result['Variable_Parameter'] = between_var
                    tukey_result['Fixed_Parameter'] = fixed_param_string
                    tukeyDf = tukeyDf.append(tukey_result.round(3))
                    
                    #Shapiro-Wilk test used to check the normal distribution of residuals
                    norm_result = pg.normality(data = df_clean[cond],
                                               dv = y_var,
                                               group = between_var,
                                               method = 'shapiro')
                    norm_result['Variable_Parameter'] = between_var
                    norm_result['Fixed_Parameter'] = fixed_param_string
                    norm_result.reset_index(inplace = True)
                    norm_result.rename(columns = {'index': 'Variable_Value'}, inplace = True)
                    normDf = normDf.append(norm_result.round(3))
                    
                    #Levene's/Bartlett’s Test to check equality of variance (levene or bartlett)
                    vareq_result = pg.homoscedasticity(data = df_clean[cond],
                                                       dv = y_var,
                                                       group = between_var,
                                                       method = 'levene')
                    vareq_result['Variable_Parameter'] = between_var
                    vareq_result['Fixed_Parameter'] = fixed_param_string
                    vareqDf = vareqDf.append(vareq_result.round(3))
                    
            else:
                tukey_result = pg.pairwise_tukey(data=df_clean,
                                                 dv= y_var,
                                                 between= between_var,
                                                 effsize = 'r')
                tukey_result['Variable_Parameter'] = between_var
                tukey_result['Fixed_Parameter'] = None
                tukeyDf = tukeyDf.append(tukey_result.round(3))
                
                #Shapiro-Wilk test used to check the normal distribution of residuals
                norm_result = pg.normality(data = df_clean,
                                           dv = y_var,
                                           group = between_var,
                                           method = 'shapiro')
                norm_result['Variable_Parameter'] = between_var
                norm_result['Fixed_Parameter'] = None
                norm_result.reset_index(inplace = True)
                norm_result.rename(columns = {'index': 'Variable_Value'}, inplace = True)
                normDf = normDf.append(norm_result.round(3))
                
                #Levene's/Bartlett’s Test to check equality of variance (levene or bartlett)
                vareq_result = pg.homoscedasticity(data = df_clean,
                                                   dv = y_var,
                                                   group = between_var,
                                                   method = 'levene')
                vareq_result['Variable_Parameter'] = between_var
                vareq_result['Fixed_Parameter'] = None
                vareqDf = vareqDf.append(vareq_result.round(3))
                
        # i = 0
        # for var in group_vars:
        #     fixed_vars = group_vars.remove(var)
        #     for f_var in fixed_vars:
                
        #     for val in exptData[fixed].unique():
        #         cond1 = df['bbh'] == 3
        #         df[cond1 & cond2]
        #         m_comp = pg.pairwise_tukey(data=df[df[fixed] == val],
        #                                 dv= y_var,
        #                                 between= fixed_params[i-1],
        #                                 effsize = 'r')
        #         m_comp['Fixed_Parameter'] = val
        #         statDf = statDf.append(m_comp).round(3)
        # ##        print(m_comp)
        #     i += 1
        # normDf.reset_index(inplace = True)
        # normDf.rename(columns = {'index': 'Variable_Value'}, inplace = True)
        # normDf.sort_values('Variable', inplace = True)
        
        vareqDf.reset_index(inplace = True)
        vareqDf.rename(columns = {'index': 'method'}, inplace = True)
        vareqDf.sort_values(['method', 'Fixed_Parameter'], inplace = True)
        
        self.statDf['anova test'] = anovaDf
        self.statDf['pairwise tukey test'] = tukeyDf
        self.statDf['normality test'] = normDf
        self.statDf['variance-equality test'] = vareqDf
        
        print(tukeyDf)
        print(normDf)
        print(vareqDf)
    
        #save statistics
        # if save == True:
        #     writer = pd.ExcelWriter("summary_statistics-" + timestamp + ".xlsx",
        #                             engine='xlsxwriter')
        #     tukeyDf.to_excel(writer, sheet_name='pairwise_t-test')
        #     aov.to_excel(writer, sheet_name='anova')
        #     norm_test.to_excel(writer, sheet_name='norm_test')
        #     var_eq_test.to_excel(writer, sheet_name='variance_eq_test')
        #     writer.save()
        
#     def plotSummary(self, summaryDict, df_filter, df_full, group = "ROI Label",
#                     marker = "o", figlist = None, leg = None):

#         if figlist == None:
#             self.figdict = {}
        
#         i = 0

# ##        header_nocomb = ["ROI Label", "Data_Folder",
# ##                         "Measurement_Number", "Measurement_OK",
# ##                         "Contact_Time", "Detachment Speed",
# ##                         "Attachment Speed", "Sliding Speed", "Sliding_Step",
# ##                         "Error_Vertical", "Error_Lateral", "Area_Units"]                      
# ##        header_comb = ["Adhesion_Force", "Adhesion_Preload",
# ##                      "Friction_Force","Friction_Preload",
# ##                      "Max_Area", "Pulloff_Area", "Friction_Area",
# ##                      "ROI_Max_Area", "ROI_Pulloff_Area",
# ##                      "Max_Length", "Pulloff_Length", "ROI_Max_Length",
# ##                      "ROI_Pulloff_Length", "Pulloff_Contact_Number",
# ##                      "Residue_Area", "Pulloff_Median_Eccentricity"]
# ##        header_all = header_nocomb + header_comb
# ##        self.df_all = pd.DataFrame(columns = header_all)

#         markerlist = ["o", "v", "P", "^", "D", "X", "<", ">", "*", "s",
#                       "+", "d", "1", "x", "2", "h"]
#         j = 0
#         print("grp", group)
#         group_unique = list(set(df_filter[group]))
#         group_unique.sort()
#         self.group_list = group_unique #if leg == None else ["All"] #only plot "All" for experiment list
# ##        self.eq_count["All"] = [1,1,1,1]
#         self.violindata = {}
#         self.violinlabels = {}
#         self.violindata["All"] = [[],[],[],[]]
#         self.violinlabels["All"] = [[],[],[],[]]
#         for b in self.group_list:
#             j = 0 if j > 15 else j #reset index
            

# ##            #combine roi data into dataframe
# ##            data_nocomb = [b] + [df_filter[x] \
# ##                                                       for x in header_nocomb \
# ##                                                       if x not in ["ROI Label"]]
# ##            data_comb = [df_filter[x + "_" + b] for x in header_comb]
# ##
# ##            df_nocomb = pd.DataFrame(dict(zip(header_nocomb, data_nocomb)))
# ##            df_comb = pd.DataFrame(dict(zip(header_comb, data_comb)))
# ##            df_joined = df_comb.join(df_nocomb)
# ##            self.df_all = self.df_all.append(df_joined, ignore_index=True, sort=False)
# ##            self.df_all['Adhesion_Force'].replace('', np.nan, inplace=True)
# ##            self.df_all.dropna(subset=['Adhesion_Force'], inplace=True) #remove blanks
            
# ##            if leg == None: #data source is summary file
#             df_roi_filter = df_filter[df_filter[group] == b]
#             roilist = [b, "All"]
# ##            else: #data source is experiment list
# ##                df_roi_filter = df_filter
# ##                roilist = [b]
                
# ##            roilist = [b, "All"] if leg == None else [b]#combine roi plots in 'All'
#             mk = markerlist[j]
#             j += 1

#             #show variable names for numeric values in legend
#             group_unit = self.get_units(group, df_roi_filter)
#             group_unit_clean = group_unit.split('(')[1].split(')')[0] if group_unit != '' else group_unit
#             self.group_name = group.replace('_', ' ') + group_unit
#             self.group_val = b
#             # if summaryDict['plot type'][0] == "Scatter":
#             b = group.replace('_', ' ') + ' ' + str(b) + group_unit_clean  \
#                 if isinstance(b, str) !=True and b!= None else b
#             leg = group.replace('_', ' ') + ' ' + str(leg) + group_unit_clean \
#                   if isinstance(leg, str) !=True and leg!= None else leg
#             # else:
#             #     b = str(b)  \
#             #         if isinstance(b, str) !=True and b!= None else b
#             #     leg = str(leg) \
#             #           if isinstance(leg, str) !=True and leg!= None else leg:
                          
# ##            if leg == None: #initialize fit equation counter
#             self.eq_count[b] = [1,1,1,1]
#             self.violindata[b] = [[],[],[],[]]
#             self.violinlabels[b] = [[],[],[],[]]
            
# ##            self.eq_count[b] = [1,1,1,1]
#             for c in roilist:
#                 c = group.replace('_', ' ') + ' ' + str(c) + group_unit_clean \
#                     if isinstance(c, str) !=True and c!= None else c

#                 # adhesion_speed_plots = {}
#                 # friction_speed_plots = {}
                
#                 title_a = summaryDict['title'][0] + ' (' + c + ')'
              
# ##                title_l = 'Adhesion (' + c + ') vs Length'
            
# ##                for a in speed_def_unique: #loop over speed definitions
# ##                    if a in ['Detachment Speed']:

#                 p1 = summaryDict['cbar var'][0] #first subplot
#                 p1_clean = p1.replace('_', ' ')
#                 p1_unit =  self.get_units(p1, df_roi_filter)
#                 x1 = summaryDict['x var'][0]
#                 x1_clean = x1.replace('_', ' ')
#                 x1_unit =  self.get_units(x1, df_roi_filter)
#                 y1 = summaryDict['y var'][0]
#                 y1_clean = y1.replace('_', ' ')
#                 y1_unit =  self.get_units(y1, df_roi_filter)
#                 title1 = 'Effect of ' + p1_clean \
#                     if summaryDict['plot type'][0] == "Scatter" else y1_clean
#                 fig_a = self.preparePlot(summaryDict['plot type'][0], 
#                                          title1, title_a, df_full, 
#                                          df_roi_filter[x1], df_roi_filter[y1],
#                                          df_roi_filter[p1], self.get_errordata(y1, df_roi_filter),
#                                          x1_clean + x1_unit,
#                                          y1_clean + y1_unit,
#                                          p1_clean + p1_unit, j,
#                                          mk if leg == None and c == "All" else marker,
#                                          figlist[c][0] if c in self.figdict.keys() else None,
#                                          b if leg == None and c == "All" else leg,
#                                          subplt = 1, fit_flag = summaryDict['fit'][0],
#                                          fit_order = summaryDict['order'][0])
# ##                adhesion_speed_plots[a] = fig_a

# ##                if a in ['Sliding Speed']:
# ##                fig_l = self.preparePlot('Effect of ' + p1_clean, title_l, df_full, 
# ##                                         df_roi_filter["Pulloff_Length"], df_roi_filter["Adhesion_Force"],
# ##                                         df_roi_filter[p1], df_roi_filter["Error_Vertical"],
# ##                                         'Contact Length ($' + df_roi_filter["Area_Units"].iloc[0][:-2] + '$)',
# ##                                         'Adhesion Force (μN)', p1_clean + ' ' + p1_unit,
# ##                                         mk if leg == None and c == "All" else marker,
# ##                                         figlist[c][1] if c in self.figdict.keys() else None,
# ##                                         b if leg == None and c == "All" else leg, subplt = 1)
# ##                friction_speed_plots[a] = fig_f

#                 p2 = summaryDict['cbar var'][1] #second subplot
#                 p2_clean = p2.replace('_', ' ')
#                 p2_unit =  self.get_units(p2, df_roi_filter)
#                 x2 = summaryDict['x var'][1]
#                 x2_clean = x2.replace('_', ' ')
#                 x2_unit =  self.get_units(x2, df_roi_filter)
#                 y2 = summaryDict['y var'][1]
#                 y2_clean = y2.replace('_', ' ')
#                 y2_unit =  self.get_units(y2, df_roi_filter)
#                 title2 = 'Effect of ' + p2_clean \
#                     if summaryDict['plot type'][0] == "Scatter" else y2_clean
#                 fig_a = self.preparePlot(summaryDict['plot type'][0], 
#                                          title2, title_a, df_full, 
#                                         df_roi_filter[x2], df_roi_filter[y2],
#                                         df_roi_filter[p2], self.get_errordata(y2, df_roi_filter),
#                                         x2_clean + x2_unit,
#                                          y2_clean + y2_unit,
#                                          p2_clean + p2_unit, j,
#                                          mk if leg == None and c == "All" else marker,
#                                         fig_a, b if leg == None and c == "All" else leg,
#                                          subplt = 2, fit_flag = summaryDict['fit'][1],
#                                          fit_order = summaryDict['order'][1])
# ##                fig_l = self.preparePlot('Effect of ' + p2_clean, title_l, df_full, 
# ##                                         df_roi_filter["Pulloff_Length"], df_roi_filter["Adhesion_Force"],
# ##                                        df_roi_filter[p2], df_roi_filter["Error_Vertical"],
# ##                                         'Contact Length ($' + df_roi_filter["Area_Units"].iloc[0][:-2] + '$)',
# ##                                         'Adhesion Force (μN)', p2_clean + ' ' + p2_unit,
# ##                                         mk if leg == None and c == "All" else marker,
# ##                                        fig_l, b if leg == None and c == "All" else leg, subplt = 2)

#                 p3 = summaryDict['cbar var'][2] #third subplot
#                 p3_clean = p3.replace('_', ' ')
#                 p3_unit =  self.get_units(p3, df_roi_filter)
#                 x3 = summaryDict['x var'][2]
#                 x3_clean = x3.replace('_', ' ')
#                 x3_unit =  self.get_units(x3, df_roi_filter)
#                 y3 = summaryDict['y var'][2]
#                 y3_clean = y3.replace('_', ' ')
#                 y3_unit =  self.get_units(y3, df_roi_filter)
#                 title3 = 'Effect of ' + p3_clean \
#                     if summaryDict['plot type'][0] == "Scatter" else y3_clean
#                 fig_a = self.preparePlot(summaryDict['plot type'][0], 
#                                          title3, title_a, df_full,  
#                                          df_roi_filter[x3], df_roi_filter[y3],
#                                         df_roi_filter[p3], self.get_errordata(y3, df_roi_filter),
#                                          x3_clean + x3_unit,
#                                          y3_clean + y3_unit,
#                                          p3_clean + p3_unit, j,
#                                          mk if leg == None and c == "All" else marker,
#                                         fig_a, b if leg == None and c == "All" else leg,
#                                          subplt = 3, fit_flag = summaryDict['fit'][2],
#                                          fit_order = summaryDict['order'][2])
# ##                fig_l = self.preparePlot('Effect of ' + p3_clean, title_l, df_full, 
# ##                                         df_roi_filter["Pulloff_Length"], df_roi_filter["Adhesion_Force"],
# ##                                        df_roi_filter[p3], df_roi_filter["Error_Vertical"],
# ##                                         'Contact Length ($' + df_roi_filter["Area_Units"].iloc[0][:-2] + '$)',
# ##                                         'Adhesion Force (μN)', p3_clean + ' ' + p3_unit,
# ##                                         mk if leg == None and c == "All" else marker,
# ##                                        fig_l, b if leg == None and c == "All" else leg, subplt = 3)

#                 p4 = summaryDict['cbar var'][3] #fourth subplot
#                 p4_clean = p4.replace('_', ' ')
#                 p4_unit =  self.get_units(p4, df_roi_filter)
#                 x4 = summaryDict['x var'][3]
#                 x4_clean = x4.replace('_', ' ')
#                 x4_unit =  self.get_units(x4, df_roi_filter)
#                 y4 = summaryDict['y var'][3]
#                 y4_clean = y4.replace('_', ' ')
#                 y4_unit =  self.get_units(y4, df_roi_filter)
#                 title4 = 'Effect of ' + p4_clean \
#                     if summaryDict['plot type'][0] == "Scatter" else y4_clean
#                 fig_a = self.preparePlot(summaryDict['plot type'][0], 
#                                          title4, title_a, df_full, 
#                                          df_roi_filter[x4], df_roi_filter[y4],
#                                         df_roi_filter[p4], self.get_errordata(y4, df_roi_filter),
#                                          x4_clean + x4_unit,
#                                          y4_clean + y4_unit,
#                                          p4_clean + p4_unit, j,
#                                          mk if leg == None and c == "All" else marker,
#                                         fig_a, b if leg == None and c == "All" else leg,
#                                          subplt = 4, fit_flag = summaryDict['fit'][3],
#                                          fit_order = summaryDict['order'][3])
# ##                fig_l = self.preparePlot('Effect of ' + p4_clean, title_l, df_full, 
# ##                                         df_roi_filter["Pulloff_Length"], df_roi_filter["Adhesion_Force"],
# ##                                        df_roi_filter[p4], df_roi_filter["Error_Vertical"],
# ##                                         'Contact Length ($' + df_roi_filter["Area_Units"].iloc[0][:-2] + '$)',
# ##                                         'Adhesion Force (μN)', p4_clean + ' ' + p4_unit,
# ##                                         mk if leg == None and c == "All" else marker,
# ##                                        fig_l, b if leg == None and c == "All" else leg, subplt = 4)

#                 self.figdict[c] = [fig_a]
                
#                 if i == 0 and c == "All" and figlist == None: #initialise figlist for "All"
#                     figlist = {}
#                     figlist["All"] = [fig_a]
#                     i = 1
# ##        self.df_final = self.df_all.copy()
# ##        self.df_all.to_excel("E:/Work/Codes/Test codes/test5.xlsx") #export as excel
                
    
#     def preparePlot(self, plot_type, ax_title, fig_title, df_full, xdata, ydata, 
#                     bardata, errdata, xlabel, ylabel, barlabel, grp_num, mk = "o", 
#                     figname = None, leg = None, subplt = None, 
#                     fit_flag = False, fit_order = 1):
#         print("preparePlot")
#         group = fig_title.split('(')[1].split(')')[0] #group value
#         ax_num = 2 if plot_type == "Scatter" else 1; #number of axis per subplot
#         if figname == None: #create figure
#             fig = plt.figure(num=fig_title, figsize = [16, 10])
#             plt.clf() #clear figure cache
# ##            fig.suptitle(fig_title, fontsize=16)
#             ax = fig.add_subplot(2,2,subplt)
#             ax.set_title(ax_title)
#             plt.cla() #clear axis cache
#             ax.set_title(ax_title)
#             # if plot_type == "Scatter":
#             ax.set_xlabel(xlabel)
#             # else:
#             #     ax.set_xlabel(self.group_name)
#             ax.set_ylabel(ylabel)
#             labels = []
#             cbar_flag = True
#         elif subplt > (len(figname.axes))/ax_num: #create subplot
#             print("a", len(figname.axes))
#             fig = figname
#             ax = fig.add_subplot(2,2,subplt)
#             ax.set_title(ax_title)
# ##            plt.cla() #clear axis cache
# ##            ax.set_title(title)
#             # if plot_type == "Scatter":
#             ax.set_xlabel(xlabel)
#             # else:
#             #     ax.set_xlabel(self.group_name)
#             ax.set_ylabel(ylabel)
#             labels = []
#             cbar_flag = True
#         else:
#             print("b", len(figname.axes))
#             fig = figname
#             ax = figname.axes[ax_num*(subplt-1)]
#             handles, labels = ax.get_legend_handles_labels()
#             cbar_flag = False
#             #increment for each new data group
#             if fit_flag == True: 
#                 self.eq_count[group][subplt-1] += 1
#         print(group, self.eq_count)
        
#         if plot_type == "Scatter":
#             if leg in labels:
#                 leg = "_nolegend_"
    
#             if bardata.dtype == 'object': #for string type data
#                 ticklabels = list(set(df_full[bardata.name]))
#                 ticklabels.sort()
#                 bardata_full = [ticklabels.index(a) for a in df_full[bardata.name]]
#                 bardata_new = [ticklabels.index(a) for a in bardata]
#                 cmin, cmax = min(bardata_full), max(bardata_full)
#             else:
#                 cmin, cmax = df_full[bardata.name].min(), df_full[bardata.name].max()
#                 ticklabels = []
#                 bardata_new = bardata
               
#             im = ax.scatter(xdata, ydata, marker = mk, s = 100, alpha = None,
#                             c = bardata_new, cmap="plasma", label = leg,
#                             vmin = cmin, vmax = cmax)
        
#             if leg != None and leg not in labels:
#                 ax.legend(loc = 'upper left')
    
#             if cbar_flag == True:
#                 print(barlabel)
#                 if bardata.dtype == 'object':
#                     cbar = fig.colorbar(im, ax = ax, ticks = [ticklabels.index(a) \
#                                                               for a in ticklabels])
#                     cbar.ax.set_yticklabels(ticklabels)
#                 else:
#                     cbar = fig.colorbar(im, ax = ax)
#                 cbar.set_clim(cmin, cmax)
#                 cbar.set_label(barlabel)
            
#             ax.errorbar(xdata, ydata,yerr= errdata,
#                         capsize = 3, ecolor = 'k', zorder=0,
#                         elinewidth = 1, linestyle="None", label = None)
    
#             if fit_flag == True:
#                 cmap = plt.cm.get_cmap('Set1')
#                 vshift = 0.05
#                 data = zip(xdata, ydata)
#                 data = np.array(sorted(data, key = lambda x: x[0]))
#                 coeff = np.polyfit(data[:,0],data[:,1], fit_order) #fitting coeffients
#                 p_fit = np.poly1d(coeff)
#                 y_fit = p_fit(data[:,0])
#                 y_avg = np.sum(data[:,1])/len(data[:,1])
#                 r2 = (np.sum((y_fit-y_avg)**2))/(np.sum((data[:,1] - y_avg)**2))
#                 sign = '' if coeff[1] < 0 else '+'
#                 eq_id = leg.split(' ')[-1] if leg != None else fig_title.split('(')[1].split(')')[0].split(' ')[-1]#[:2]
#                 eq_coff = ["$%.1e"%(coeff[i]) + "x^" + str(len(coeff) - i - 1) + "$"\
#                      if i < len(coeff) - 2 else "%.4fx"%(coeff[i]) for i in range(len(coeff)-1)]
#                 eq =  "y=" + '+'.join(eq_coff) + "+%.4f"%(coeff[len(coeff)-1]) + "; $R^2$=" + "%.4f"%(r2)  
#                 eq_clean = eq.replace('+-', '-')
#                 x_fit = np.linspace(min(data[:,0]), max(data[:,0]), 100)
#                 ax.plot(x_fit, p_fit(x_fit), color = cmap(self.eq_count[group][subplt-1]*0.1),
#                         linewidth=1, linestyle='dotted')
#                 ax.text(1,0.2 - (vshift * self.eq_count[group][subplt-1]),
#                         eq_id + ": " + eq_clean, ha = 'right',
#                         transform=ax.transAxes, color = cmap(self.eq_count[group][subplt-1]*0.1),
#                         bbox=dict(facecolor='white', edgecolor = 'white', alpha=0.5))
#     ##            self.eq_count[subplt-1] += 1
        
#         elif plot_type in ["Box","Violin"]:                      
#             print("Box",leg)
#             ax.cla()
#             ax.set_ylabel(ylabel)
#             self.violindata[group][subplt-1].append(ydata)
#             datasize = str(len(ydata))
#             if group == "All":
#                 # group_size = len(self.group_list)
#                 # boxdata = [[]]*group_size
#                 # boxdata[grp_num-1] = ydata
#                 # boxlabels = self.group_list
#                 # # boxlabels = [[]]*group_size
#                 # boxlabels[grp_num-1] = leg
#                 # boxpositions = list(range(1,group_size+1))
#                 self.violinlabels[group][subplt-1].append(str(self.group_val) + 
#                                                           '\n' + '(n=' + 
#                                                           datasize + ')')
#                 ax.set_xlabel(self.group_name)
#             else:
#                 # boxdata = ydata
#                 # boxlabels = [group]
#                 # boxpositions = [1]
#                 self.violinlabels[group][subplt-1].append(str(group) + '\n' + 
#                                                           '(n=' + datasize + 
#                                                           ')')
#                 ax.set_xlabel(None)
#             violinpositions = list(range(1,len(self.violinlabels[group][subplt-1])+1)) 
#             if plot_type == "Box":
#                 ax.boxplot(self.violindata[group][subplt-1], positions=violinpositions,
#                            labels=self.violinlabels[group][subplt-1])
#             elif plot_type == "Violin":
#                 # self.violindata[group][subplt-1].append(ydata)
#                 # self.violinlabels[group][subplt-1].append(leg)
#                 # violinpositions = list(range(1,len(self.violinlabels[group][subplt-1])+1))
#                 ax.violinplot(self.violindata[group][subplt-1], positions=violinpositions, 
#                               showmedians=True)
#                 ax.set_xticks(violinpositions)
#                 ax.set_xticklabels(self.violinlabels[group][subplt-1])

#         fig.tight_layout()
# ##        fig.show()
# ##        plt.show()
        
#         return fig

    def showSummaryPlot(self): #show summary plots
        print("showSummaryPlot")
        # if self.summary_filepath != "":
#             keys = list(self.figdict.keys())
#             for b in keys:
#                 print("keys", b)
#                 if len(self.figdict.keys())==2 and b == "All":
#                     #close "All" figures
#                     plt.close(self.figdict[b][0])
# ##                    plt.close(self.figdict[b][1])
# ##                    plt.close(self.figdict[b][2])
# ##                    plt.close(self.figdict[b][3])
# ##                    plt.close(self.figdict[b][4])
# ##                    plt.close(self.figdict[b][5])
# ##                    for a in self.figdict[b][6].values():
# ##                        plt.close(a)
# ##                    for a in self.figdict[b][7].values():
# ##                        plt.close(a)
#                 else:
# ##                    self.figdict[b][0].show()
#                     self.show_figure(self.figdict[b][0])
# ##                    self.figdict[b][1].show()
# ##                    self.figdict[b][2].show()
# ##                    self.figdict[b][3].show()
# ##                    self.figdict[b][4].show()
# ##                    self.figdict[b][5].show()
# ##                    for a in self.figdict[b][6].values():
# ##                        a.show()
# ##                    for a in self.figdict[b][7].values():
# ##                        a.show()
        self.previewPlot()
        self.fig.fig.show()
        plt.show()

    # def show_figure(self, fig):
    #     # create a dummy figure and use its
    #     # manager to display "fig"
    #     dummy = plt.figure(num=fig.get_label(), figsize = [16, 10])
    #     new_manager = dummy.canvas.manager
    #     new_manager.canvas.figure = fig
    #     fig.set_canvas(new_manager.canvas)    
    
    #save summary plots
    def saveSummaryPlot(self, folderpath, plot_format): 
        # if self.summary_filepath != "":
            # folderpath = os.path.dirname(self.summary_filepath)
            # if not os.path.exists(folderpath):orientation='landscape',
            #     os.makedirs(folderpath)
        filename = folderpath + '/' + \
            self.fig.fig.canvas.get_window_title()+ '-' + \
                time.strftime("%y%m%d%H%M%S") + '.' + plot_format
        self.fig.fig.savefig(filename, bbox_inches='tight', 
                             transparent = True, dpi = 150)
            # keys = list(self.figdict.keys())
            # for b in keys:
            #     if len(self.figdict.keys())==2 and b == "All":
            #         continue
            #     else:
            #         self.savePlot(self.figdict[b][0], plot_format)
##                    self.savePlot(self.figdict[b][1])
##                    self.savePlot(self.figdict[b][2])
##                    self.savePlot(self.figdict[b][3])
##                    self.savePlot(self.figdict[b][4])
##                    self.savePlot(self.figdict[b][5])
##                    for a in self.figdict[b][6].values():
##                        self.savePlot(a)
##                    for a in self.figdict[b][7].values():
##                        self.savePlot(a)
        print("saved plot", filename)
        # self.summarydf.to_excel(os.path.dirname(summary_filepath) +
        #                        '/summary_clean_' +
        #                        time.strftime("%y%m%d%H%M%S") + '.xlsx') #export as excel 

        # print("saved data")
            
    # def savePlot(self, fig, plot_format): #save routine
    #     filename = os.path.dirname(self.summary_filepath) + '/' + \
    #                fig.get_label().replace('/','')+ '-' + time.strftime("%y%m%d%H%M%S") + '.' + plot_format
    #     # fig.savefig(filename, orientation='landscape',
    #     #             transparent = True, dpi = 150)

    #     print("save plot", filename)
    
    #combine summary data from experiment list
    def combineSummaryFromList(self, list_filepath, subfolder): 
        # root = tk.Tk()
        # root.withdraw()
        # self.list_filepath, _ =  QFileDialog.getOpenFileName(caption = 
        #                                                      "Select experiment list file")
        # if self.list_filepath != "":
        list_folderpath = os.path.dirname(list_filepath)
        
        explistDf = pd.read_excel(list_filepath) #experiment list
        
        col_list = explistDf.columns
        listunitDict = {}
        for col_name in col_list:
            unit = col_name.split('(')[-1].split(')')[0]
            if unit == col_name:
                listunitDict[col_name] = ''
            else:
                col_clean = col_name.split('(')[0].strip()
                explistDf.rename(columns = {col_name : col_clean}, 
                                 inplace=True)
                listunitDict[col_clean] = ' [$' + unit + '$]'
        
        # folder_heirarchy = '/Analysis/Summary/'
        
        fullDf = None
        
        for i in explistDf.index:
            exp_data = explistDf.loc[i]
            if exp_data['Data OK?'] == 'Yes' and exp_data['Include Data?'] == 'Yes':
                summary_folder = list_folderpath + "/" + \
                    exp_data['Data Folder name'] + subfolder
                with os.scandir(summary_folder) as folder:
                    for file in folder:
                        if file.name.endswith('.txt') and file.is_file():
                            summarydf = self.importSummary(file.path)
                            joinedDf = summarydf.merge(
                                pd.DataFrame(data = [exp_data.values]*len(exp_data),
                                             columns = exp_data.index),
                                left_index=True, right_index=True)
                            if fullDf.__class__.__name__ == 'NoneType':
                                fullDf = joinedDf.copy()
                            else:
                                fullDf = fullDf.append(joinedDf, ignore_index=True, 
                                                       sort=False)
        
        self.unitDict = {**listunitDict, **self.unitDict}
        
        return fullDf
    
    #combine text files from same folder
    def combineSummaryFromFolder(self, folderpath, subfolder):
        fullDf = None        
        with os.scandir(folderpath + '/') as folder:
            for file in folder:
                if file.name.endswith('.txt') and file.is_file() and subfolder == '':
                    summarydf = self.importSummary(file.path)
                    if fullDf.__class__.__name__ == 'NoneType':
                        fullDf = summarydf.copy()
                    else:
                        fullDf = fullDf.append(summarydf, ignore_index=True, 
                                               sort=False)
                elif file.is_dir() and subfolder != '':
                    dir_path = folderpath + '/' + file.name + '/'
                    with os.scandir(dir_path) as dir_inside:
                        for file_in in dir_inside:
                            if file_in.name.endswith('.txt') and file_in.is_file():
                                summarydf = self.importSummary(file_in.path)
                                if fullDf.__class__.__name__ == 'NoneType':
                                    fullDf = summarydf.copy()
                                else:
                                    fullDf = fullDf.append(summarydf, ignore_index=True, 
                                                           sort=False)                    
        
        return fullDf
    
                        
                
#             wb_obj = openpyxl.load_workbook(filename = self.list_filepath,
#                                             read_only = True)# workbook object is created  
#             sheet_obj = wb_obj.active

#             m_row = sheet_obj.max_row

#             date = []
#             foldername = []
#             species = []
#             sex = []
#             leg = []
#             pad = []
#             weight = []
#             temp = []
#             hum = []
#             medium = []
#             surface = []
#             ca_w = []
#             ca_o = []
#             dataok = []
#             includedata = []
#             label = []
            
#             header1 = ["Date", "Folder_Name", "Species", "Sex", "Leg", "Pad",
#                        "Weight", "Temperature", "Humidity", "Medium",
#                        "Substrate","Contact_Angle-Water", "Contact_Angle-Hexadecane", 
#                        "Data_OK", "Include_Data", "Label"]
# ##            header2 = ["Max_Area", "Pulloff_Area","Adhesion_Force",
# ##                      "Preload_Force", "Contact_Time", "Speed",
# ##                      "Steps", "Friction_Force", "Friction_Area",
# ##                      "Measurement_Number", "Measurement_OK", "ROI_Labels",
# ##                      "Area_Units"]
# ##            headerfull = header1 + header2
#             df = pd.DataFrame(columns = header1)

# ##            steps_unique = []
# ##            speed_def_unique = []
# ##            roi_label_unique = []
#             j = 0
            
#             for i in range(3, m_row + 1):
#                 #import data
#                 ok = sheet_obj.cell(row = i, column = 16).value
#                 include = sheet_obj.cell(row = i, column = 17).value

#                 if ok == 'No' or include == 'No': #only consider 'Yes' data in Data OK/Include Data
#                     continue
                
#                 date.append(sheet_obj.cell(row = i, column = 1).value)
#                 foldername.append(sheet_obj.cell(row = i, column = 2).value)
#                 species.append(sheet_obj.cell(row = i, column = 3).value)
#                 sex.append(sheet_obj.cell(row = i, column = 4).value)
#                 leg.append(sheet_obj.cell(row = i, column = 5).value)
#                 pad.append(sheet_obj.cell(row = i, column = 6).value)
#                 weight.append(sheet_obj.cell(row = i, column = 7).value)
#                 temp.append(sheet_obj.cell(row = i, column = 8).value)
#                 hum.append(sheet_obj.cell(row = i, column = 9).value)
#                 medium.append(sheet_obj.cell(row = i, column = 10).value)
#                 surface.append(sheet_obj.cell(row = i, column = 11).value)
#                 ca_w.append(sheet_obj.cell(row = i, column = 12).value)
#                 ca_o.append(sheet_obj.cell(row = i, column = 13).value)
#                 dataok.append(ok)
#                 includedata.append(include)
#                 label.append(sheet_obj.cell(row = i, column = 18).value)

#                 print(foldername[j], m_row)
#                 if foldername[j] != None:
#                     self.importSummary(list_folderpath + "/" + foldername[j] +
#                                   "/Analysis/Summary/summary data.txt")
# ##                    steps_unique.append(self.steps_unique)
# ##                    roi_label_unique.append(self.roi_label_unique)
# ##                    roi_label_unique.append(set(self.df_forcedata["ROI Label"]))
# ##                    speed_def_unique.append(self.speed_def_unique)
#                     rownum = len(self.df_forcedata["Max_Area"])
#                     datalist = [[date[j]]*rownum, [foldername[j]]*rownum, 
#                                 [species[j]]*rownum,[sex[j]]*rownum, 
#                                 [leg[j]]*rownum, [pad[j]]*rownum, 
#                                 [weight[j]]*rownum,[temp[j]]*rownum, 
#                                 [hum[j]]*rownum, [medium[j]]*rownum, 
#                                 [surface[j]]*rownum, [ca_w[j]]*rownum,
#                                 [ca_o[j]]*rownum, [dataok[j]]*rownum, 
#                                 [includedata[j]]*rownum, [label[j]]*rownum]
#                     datadict = dict(zip(header1, datalist))
#                     df_data = pd.DataFrame(datadict)
#                     df_joined = df_data.join(self.df_forcedata)
#                     df = df.append(df_joined, ignore_index=True, sort=False)
# ##                    df.to_excel('E:/Work/Data/Summary/20200213/Sex/summary_comb_' +
# ##                                str(random.randint(1, 90000)) + '.xlsx') #export as excel                    
# ##                    print(df.to_string())
#                 else:
#                     break
#                 j += 1
            
#             wb_obj.close()
#             print("import finish")

#             # df['Date'] = df['Date'].dt.strftime('%d/%m/%Y')
#             df['Date'] = pd.to_datetime(df['Date'], format = '%d=%m-%Y').dt.date

# ##            roi_label_unique = list(set([a for b in roi_label_unique for a in b]))
# ##            speed_def_unique = list(set([a for b in speed_def_unique for a in b]))


#             self.df_final = df.copy()

            # self.df_final = fullDf.copy()

            #save summary combined
                
            # excel_folderpath = list_folderpath + '/Summary/' +  \
            #                  time.strftime("%Y%m%d") + '/' + legend_parameter
            # excel_filepath = excel_folderpath + '/summary_combined_' +  \
            #                  time.strftime("%Y%m%d%H%M%S") + '.xlsx'
            
##            if not os.path.exists(excel_folderpath):
##                os.makedirs(excel_folderpath)
            
##            self.df_all.to_excel(excel_filepath) #export as excel

            # df_good = self.df_final
            # self.summary_filepath = 'CHECK' #to save plots in Summary directory
            # self.plotSummary(summaryDict,
            #                  df_good,
            #                  df_good,
            #                  legend_parameter)            

    #         if legend_parameter == "ROI Label": #no filtering as this is already plotted in prepareplot (when leg = None)          
    #             self.plotSummary(summaryDict, df_good, df_good)
    #         else:
    #         ##    legend_parameter = 'Folder_Name' #choose, same as column names
    #             legend_list = df_good[legend_parameter].unique()
    #             legend_list.sort()
    #             print(legend_list)
    #             markerlist = ["o", "v", "P", "^", "D", "X", "<", ">", "*", "s",
    #                           "+", "d", "1", "x", "2", "h"]
    #             figlist = None
    #             i = 0
    # ##            df_leg = pd.DataFrame(dict(zip([legend_parameter], [legend_list])))
    #             for lg in legend_list:
    #                 print("zxz", lg)
    #                 i = 0 if i > 15 else i
    #                 df_filtered = df_good[df_good[legend_parameter] == lg]
    #                 self.plotSummary(summaryDict,
    #                                  df_filtered, df_good, legend_parameter, markerlist[i],
    #                                  figlist, lg)
    #                 figlist = self.figdict.copy()
    # ##                df_all_joined = self.df_all.copy()
    # ##                df_all_joined.insert(0, legend_parameter, lg)
    # ##                if i == 0:
    # ##                    df_final = df_all_joined.copy()
    # ##                else:
    # ##                    df_final = df_final.append(df_all_joined, ignore_index=True, sort=False)
    # ##                print("iter", i)
    #                 i += 1
    # ##            self.df_final = df_final.copy() 
            

##a.combineSummary("Folder_Name")
##if a.list_filepath != "":
##    a.showSummaryPlot()
##summary = SummaryAnal()
##summary.importSummary()
##summary.plotSummary(summary.speed_def_unique,
##                    summary.roi_label_unique,
##                    summary.df_forcedata,
##                    summary.df_forcedata)
##summary.showSummaryPlot()
        
