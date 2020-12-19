# -*- coding: utf-8 -*-
"""
Created on Fri Jun 19 12:10:12 2020

@author: adwait
"""
import pims
import numpy as np
import cv2
import matplotlib.pyplot as plt
import openpyxl
import os.path
import logging
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtChart import QScatterSeries
# from source.analysis.forceanalysis import ForceAnal
from source.threads.countframethread import CountFrameThread

class MainImportFile:

    def load_image(self): #load image file
        self.videoPath, _ = QFileDialog.getOpenFileName(self, "Open Image File")
        if self.videoPath != "":
            self.cap = pims.open(self.videoPath)
            # self.frame = cv2.imread(self.videoPath)
            self.frame = self.cap[0]
            self.playStatus = False
            self.recordStatus = False
            self.frameCount = 1
            self.frameHeight, self.frameWidth = self.frame.shape[:2]
            roiCorners = np.array([[0, 0],[self.frameWidth, 0], 
                                        [self.frameWidth, self.frameHeight], 
                                        [0, self.frameHeight]],np.int32)
            self.roiBound = [0, 0, self.frameWidth, self.frameHeight]
            self.roiDict = {"Default": [roiCorners, self.roiBound, [], roiCorners, roiCorners]}
            self.frameBackground = 255 * np.ones((self.frameHeight,
                                                self.frameWidth, 3),dtype=np.uint8)
            self.renderVideo("Raw", self.frame)
            self.frame_current = self.frame.copy()
            #Frame no. ROI label, contour id, area, length, ecc, array
            self.contour_data = [[], [], [], [], [], [], [], []] 
            # contactArea = np.zeros(1, np.float64)
            # contactLength = np.zeros(1, np.float64)
            # contourNumber = np.zeros(1, np.uint64)
            # roiArea = np.zeros(1, np.float64)
            # roiLength = np.zeros(1, np.float64)
            # self.frameTime = np.zeros(1, np.float64)
            # eccAvg = np.zeros(1, np.float64)
            # contactAngle =  np.zeros(1, np.float64)
            # self.dataDict = {"Default" : {"Contact area": contactArea, 
            #                               "Contact length": contactLength,
            #                               "Contact number": contourNumber, 
            #                               "ROI area": roiArea,
            #                               "ROI length": roiLength, 
            #                               "Eccentricity": eccAvg,
            #                               "Ellipse fit": [(0,0),(0,0),0,1],
            #                               "Contact angle": contactAngle}
            #                  }
            self.dataDict = {}
            self.init_datadict("Default") #initialize self.dataDict
            self.effectChain = [True, False, False, False] #b/c, hist, bg sub, filter
##            self.roi_auto = self.threshROIGroupBox.isChecked()
            self.distinct_roi = self.distinctAutoROI.isChecked()
            self.roi_hull = self.applyHullROI.isChecked()
            self.combine_roi = self.combineROI.isChecked()
            self.roi_tresh_type = "Global"
            self.tresh_size_roi = self.threshROISpinBox1.value()
            self.tresh_cst_roi = self.threshROISpinBox2.value()
            self.blur_size_roi = self.blurROISpinBox.value()
            self.bg_roi_apply = self.applybgROI.isChecked()
            self.bg_blur_size_roi = self.bgblurROISpinBox.value()
            self.bg_blend_roi = self.bgblendROISpinBox.value()
            self.framePos = 0
            self.roi_min_area = self.roiMinSpinBox.value()
            self.epsilon_fraction = 10**(self.epsilonSpinBox.value())
            self.roi_morph = self.morphROI.isChecked()
            self.x_roi_morph = self.morphXSpinBox.value()
            self.y_roi_morph = self.morphYSpinBox.value()
            self.calibFactor = self.lengthValue.value()/self.pixelValue.value()
            
            self.definePaths()
            
            self.analyzeVideo.blockSignals(True)
            self.analyzeVideo.setChecked(False)
            self.analyzeVideo.blockSignals(False)
            
            self.dftGroupBox.blockSignals(True)
            self.dftGroupBox.setChecked(False)
            self.dftGroupBox.blockSignals(False)

            self.bgGroupBox.blockSignals(True)
            self.bgGroupBox.setChecked(False)
            self.bgGroupBox.blockSignals(False)

            # self.showContours.blockSignals(True)
            # self.showContours.setChecked(False)
            # self.showContours.blockSignals(False)
            self.rawViewTab.setTabEnabled(1,False)

            self.threshROIGroupBox.blockSignals(True)
            self.threshROIGroupBox.setChecked(False)
            self.threshROIGroupBox.blockSignals(False)

##            self.distinctAutoROI.setEnabled(False)
            self.roi_auto = False
            self.segment = self.segmentGroupBox.isChecked()
            self.show_segment = self.showSegment.isChecked()
            self.show_fg = self.segmentFGButton.isChecked()
            self.show_bg = self.segmentBGButton.isChecked()
##            self.distinct_roi = False
            
##            self.distinctAutoROI.blockSignals(True)
##            self.distinctAutoROI.setChecked(False)
##            self.distinctAutoROI.blockSignals(False)

            # self.correctZeroForce.blockSignals(True)
            # self.correctZeroForce.setChecked(False)
            # self.correctZeroForce.blockSignals(False)

            self.analyzeDataWindow.zero_subtract.setChecked(False)     
            self.analyzeDataWindow.zero_subtract.setEnabled(False)

            
            self.init_dict() #initialise dictionaries
                        
            # self.forceData = ForceAnal(self.fitWindow, self.configPlotWindow,
            #                            self.analyzeDataWindow)
            # self.zeroForceData = ForceAnal() #initalise
            # self.correctZeroForce.setEnabled(False)
            self.clearData.setEnabled(True)
            # self.videoEffect.setCurrentIndex(0) CHECK TAB DISABLE!
            # self.videoEffect.model().item(5).setEnabled(False)
            # self.videoEffect.model().item(6).setEnabled(True)
            self.effectViewTab.setTabEnabled(4,False)
            # self.effectViewTab.setTabEnabled(6,True)
            
            self.videoFileNameLabel.setText("<b>Video file:</b>\n" + self.videoPath) #video path
            # self.forceFileNameLabel.setText("Select force data from file menu")
            # self.zeroForceFileNameLabel.setText("Select zero-force data from file menu")
            self.statusBar.showMessage("Frame number: " + str(int(self.framePos)))
            
            self.frame_bin_roi_full = None
            self.setWindowTitle(self.appVersion)
            self.msrmnt_num = None
            self.bgframeNumber = 0
           
    def load_video(self): #load video
        if self.msrListMode == False:
            self.videoPath, _ = QFileDialog.getOpenFileName(self, "Open Video File")
        if self.videoPath != "":
##            self.statusBar.showMessage("Loading...")
            # self.cap = cv2.VideoCapture(self.videoPath)
            self.cap = pims.Video(self.videoPath)
            self.playStatus = False
            self.frameAction = "Play"
            self.recordStatus = False
            # self.setVideoParam()
##            self.frameCount = self.cap.get(cv2.CAP_PROP_FRAME_COUNT)
            #BUG: frame number count calculation is dirty. fix it.
            self.countThread = CountFrameThread(cv2.VideoCapture(self.videoPath)) #count frames in separate thread
            self.countThread.output.connect(self.loading_indicate)
            self.countThread.finished.connect(self.setVideoParam)
            self.countThread.start()

    def setVideoParam(self): #set video parameters
##            self.count_frames() #count number of frames by loop
            # self.cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
            self.frameCount = self.countThread.frameCount
            # self.frameWidth = int(self.cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            # self.frameHeight = int(self.cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
   
            # self.frameCount = len(self.cap)
            self.frameHeight, self.frameWidth, _ = self.cap.frame_shape
            self.framePos = 1
            try:
                self.fpsSpinBox.setValue(self.cap.frame_rate)
            except Exception as e:
                logging.warning(f'{e}')   
            logging.debug('%s, %s, %s', self.frameCount, self.frameWidth, self.frameRate)
            
            roiCorners = np.array([[0, 0],[self.frameWidth, 0], 
                                        [self.frameWidth, self.frameHeight], 
                                        [0, self.frameHeight]],np.int32)
            self.roiBound = [0, 0, self.frameWidth, self.frameHeight]
            #roi dictionary: actual roi, bound, contours, adjusted roi, auto roi
            self.roiDict = {"Default": [roiCorners, self.roiBound,
                                        [], roiCorners, roiCorners]} 
            self.frameBackground = 255 * np.ones((self.frameHeight,
                                                self.frameWidth, 3),dtype=np.uint8)
            self.bgframeNumber = None
            
            # self.ret, self.frame = self.cap.read()
            # if self.ret == False: #reset video on error
            #         logging.debug("if")
            #         self.cap.release()
            #         self.cap = cv2.VideoCapture(self.videoPath)
            #         self.ret, self.frame = self.cap.read()
            self.frame = self.cap[0]
            self.renderVideo("Raw", self.frame)

            self.effectScene.removeItem(self.effectPixmapItem)
            self.effectPixmapItem = self.effectScene.addPixmap(self.blankPixmap)
            self.effectView.fitInView(self.effectPixmapItem, 1)
                
            # self.cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
            self.frame_current = self.frame.copy()
            #Frame no. ROI label, contour id, area, length, ecc, array
            self.contour_data = [[], [], [], [], [], [], [], []]
            # contactArea = np.zeros(int(self.frameCount), np.float64)
            # contactLength = np.zeros(int(self.frameCount), np.float64)
            # contourNumber = np.zeros(int(self.frameCount), np.uint64)
            # roiArea = np.zeros(int(self.frameCount), np.float64)
            # roiLength = np.zeros(int(self.frameCount), np.float64)
            # eccAvg = np.zeros(int(self.frameCount), np.float64) #median eccentricity
            # contactAngle = np.zeros(int(self.frameCount), np.float64)
            # self.dataDict = {"Default" : {"Contact area": contactArea, 
            #                               "Contact length": contactLength,
            #                               "Contact number": contourNumber, 
            #                               "ROI area": roiArea,
            #                               "ROI length": roiLength, 
            #                               "Eccentricity": eccAvg,
            #                               "Ellipse fit": [[(0,0),(0,0),0,1]]*int(self.frameCount),
            #                               "Contact angle": contactAngle}
            #                  }
            self.dataDict = {}
            self.init_datadict("Default") #initialize self.dataDict
            self.frameTime = np.linspace(0,
                                         self.frameCount/self.frameRate,
                                         int(self.frameCount), dtype = np.float64)
            self.effectChain = [True, False, False, False] #b/c, hist, bg sub, filter
##            self.roi_auto = self.threshROIGroupBox.isChecked()
            self.distinct_roi = self.distinctAutoROI.isChecked()
            self.roi_hull = self.applyHullROI.isChecked()
            self.combine_roi = self.combineROI.isChecked()
            self.roi_tresh_type = "Global"
            self.tresh_size_roi = self.threshROISpinBox1.value()
            self.tresh_cst_roi = self.threshROISpinBox2.value()
            self.blur_size_roi = self.blurROISpinBox.value()
            self.bg_roi_apply = self.applybgROI.isChecked()
            self.bg_blur_size_roi = self.bgblurROISpinBox.value()
            self.bg_blend_roi = self.bgblendROISpinBox.value()
            # self.framePos = 0
            self.roi_min_area = self.roiMinSpinBox.value()
            self.epsilon_fraction = 10**(self.epsilonSpinBox.value())
            self.roi_morph = self.morphROI.isChecked()
            self.x_roi_morph = self.morphXSpinBox.value()
            self.y_roi_morph = self.morphYSpinBox.value()
            self.calibFactor = self.lengthValue.value()/self.pixelValue.value()

            self.definePaths()
            
            self.seekSlider.blockSignals(True)
            self.seekSlider.setValue(1)
            self.seekSlider.setTickInterval(int(0.2 *self.frameCount))
            self.seekSlider.setSingleStep(int(0.1 *self.frameCount))
            self.seekSlider.setMaximum(int(self.frameCount))
            self.seekSlider.blockSignals(False)
            
            self.analyzeVideo.blockSignals(True)
            self.analyzeVideo.setChecked(False)
            self.analyzeVideo.blockSignals(False)
            
            self.dftGroupBox.blockSignals(True)
            self.dftGroupBox.setChecked(False)
            self.dftGroupBox.blockSignals(False)

            self.bgGroupBox.blockSignals(True)
            self.bgGroupBox.setChecked(False)
            self.bgGroupBox.blockSignals(False)

            self.threshROIGroupBox.blockSignals(True)
            self.threshROIGroupBox.setChecked(False)
            self.threshROIGroupBox.blockSignals(False)

##            self.distinctAutoROI.setEnabled(False)
            self.roi_auto = False
            self.segment = self.segmentGroupBox.isChecked()
            self.show_segment = self.showSegment.isChecked()
            self.show_fg = self.segmentFGButton.isChecked()
            self.show_bg = self.segmentBGButton.isChecked()
##            self.distinct_roi = False
            
##            self.distinctAutoROI.blockSignals(True)
##            self.distinctAutoROI.setChecked(False)
##            self.distinctAutoROI.blockSignals(False)

            # self.correctZeroForce.blockSignals(True)
            # self.correctZeroForce.setChecked(False)
            # self.correctZeroForce.blockSignals(False)
            
            self.analyzeDataWindow.zero_subtract.setChecked(False)     
            self.analyzeDataWindow.zero_subtract.setEnabled(False)
            

            self.rawViewTab.setTabEnabled(1,False)
            # self.showContours.setEnabled(False)
            # self.showEffect.setEnabled(False)
            
            
            # self.forceData = ForceAnal(self.fitWindow, self.configPlotWindow,
            #                            self.analyzeDataWindow)
            # self.zeroForceData = ForceAnal() #initalise
            # self.correctZeroForce.setEnabled(False)
            self.clearData.setEnabled(True)
            # self.videoEffect.setCurrentIndex(0) CHECK TAB DISABLE!
            # self.videoEffect.model().item(5).setEnabled(False)
            # self.videoEffect.model().item(6).setEnabled(True)
            self.effectViewTab.setTabEnabled(4,False)
            # self.effectViewTab.setTabEnabled(6,True)

            self.videoFileNameLabel.setText("<b>Video file:</b>\n" + self.videoPath) #video path
            # self.forceFileNameLabel.setText("Select force data from file menu")
            # self.zeroForceFileNameLabel.setText("Select zero-force data from file menu")
##            self.statusBar.showMessage("Frame number: " + str(int(self.framePos)) + "\t("
##                                       + str(int((self.framePos)*100/self.frameCount)) +
##                                       "%)")
            self.statusBar.showMessage(str(int(self.frameCount)) + " frames loaded!")
            self.configRecWindow.videoTextbox.setText("")

            self.init_dict() #initialise dictionaries
            
            # self.framePos = 1 #avoid array indexing issue
            self.frame_bin_roi_full = None

            self.curve1 = QScatterSeries()#initialise live plot curves
            self.curve2 = QScatterSeries()
            self.initialise_live_plot(self.curve1, Qt.blue) #contact area plot
            self.initialise_live_plot(self.curve2, Qt.red) # roi area plot

            self.msrmnt_num = None
            self.setWindowTitle(self.appVersion)

            if self.msrListMode == True: #continue loading force data etc
                self.load_measurement_data()

    def import_file_list(self): #import file list
        fileListPath, _ = QFileDialog.getOpenFileName(self, "Open File List")

        if fileListPath != "":
            self.folderPath = os.path.dirname(fileListPath)
            wb_obj = openpyxl.load_workbook(filename = fileListPath,
                                            read_only = True)# workbook object is created  
            sheet_obj = wb_obj.active

            m_row = sheet_obj.max_row

            self.measurmntList = []
            self.bottomviewList = []
            self.sideviewList = []
            self.forcedataList = []
             
            for i in range(2, m_row + 1): #save filename lists
                self.measurmntList.append(sheet_obj.cell(row = i, column = 1).value)
                self.bottomviewList.append(sheet_obj.cell(row = i, column = 2).value)
                self.sideviewList.append(sheet_obj.cell(row = i, column = 3).value)
                self.forcedataList.append(sheet_obj.cell(row = i, column = 4).value)

            wb_obj.close()
            self.chooseMsrmnt.setEnabled(True)
            self.statusBar.showMessage("Loaded file list from " + self.folderPath)
            logging.debug('%s, %s', self.measurmntList, self.folderPath)

    def import_force_data(self): #import force data
        # if self.msrListMode == False:
        #     self.forceData = ForceAnal(self.fitWindow, self.configPlotWindow,
        #                                self.analyzeDataWindow)
        # self.zeroForceData = ForceAnal() #initalise
        # self.correctZeroForce.blockSignals(True)
        # self.correctZeroForce.setChecked(False)
        # self.correctZeroForce.blockSignals(False)        
        # self.correctZeroForce.setEnabled(False)
        self.analyzeDataWindow.zero_subtract.setChecked(False)     
        self.analyzeDataWindow.zero_subtract.setEnabled(False)
        self.zeroForceFileNameLabel.setText("Select zero-force data from file menu")
        
        self.forceData.importData(self.msrListMode)
        if self.forceData.force_filepath != "":
            logging.info('File loaded -> ' + self.forceData.force_filepath)
            self.forceData.calcData()
            # self.defl_vert1_raw = self.forceData.defl_vert1.copy() #copy of raw vert data
            # self.defl_vert1_actual = self.forceData.defl_vert1.copy() #copy of raw vert data
            self.configPlotWindow.plotDict['plot settings']['plot range']\
                .setText(str(0) + ',' + str(self.forceData.ptsnumber-1))
            # self.forceData.plot_slice = self.configPlotWindow.plotDict['plot settings']['plot range']
            self.openZeroForceFile.setEnabled(True)
            self.fpsSpinBox.blockSignals(True)
            self.fpsSpinBox.setValue(self.forceData.fps)
            self.fpsSpinBox.blockSignals(False)
            self.frameTime = self.forceData.time_video #recalculated time array
            self.frameRate = self.forceData.fps
            self.forceFileNameLabel.setText("<b>Force data:</b>\n" + self.forceData.force_filepath)
##            self.videoEffect.model().item(6).setEnabled(True)
            plt.close()
            self.plotSequence()
            # self.updateFitDict()
            # self.init_plotconfig()
            # self.forceData.dataClean()
            # self.forceData.calcData()
            logging.debug('%s', self.frameRate)

    def import_zero_force(self): #import zero force line
        # self.zeroForceData = ForceAnal(analyzeDataWindow = self.analyzeDataWindow)
        # self.correctZeroForce.setEnabled(True)
        self.analyzeDataWindow.zero_subtract.setEnabled(True)
        self.analyzeDataWindow.zero_subtract.setChecked(False)             
        # self.zeroForceData.noiseSteps = ','.join(map(str,range(1,self.forceData.step_num + 1)))
        self.zeroForceData.importData(False)
        if self.zeroForceData.force_filepath != "":
            self.zeroForceData.dataClean()
            self.zeroForceData.evaluateForce()
            self.forceData.zeroDataDict = self.zeroForceData.fileDataDict.copy()
            self.zeroForceFileNameLabel.setText("<b>Zero force data:</b>\n" +
                                                self.zeroForceData.force_filepath)            
            # self.zero_force_calc()

#     def zero_force_calc(self): #calculate vertical force correction wrt zero
#         zero_shift = self.zeroForceData.defl_vert1[0] - self.forceData.defl_vert1[0]
#         self.zeroForceData.defl_vert1_corrected = [x-zero_shift for x in self.zeroForceData.defl_vert1]
# ##        self.defl_vert1_raw = self.forceData.defl_vert1.copy()
#         self.defl_vert1_actual = [self.forceData.defl_vert1[0] +
#                                  self.forceData.defl_vert1[i] -
#                                  self.zeroForceData.defl_vert1_corrected[i] \
#                                  for i in range(len(self.forceData.defl_vert1))]